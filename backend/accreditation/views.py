import os

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseRedirect, FileResponse
from django.conf import settings
from django.urls import reverse
from django.db.models import Q
import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
import openpyxl
from .database import *
from openpyxl.workbook import Workbook
import json
from .utils import *

# Import models [NEEDS TO BE CHANGED/DELETED]
from .models import (
    DataProcess,
    FacultyCI,
    ProgramCI,
    AssessValidity,
    AccredReport,
    AnnualReport,
    Faculty,
)


# Check if user is admin
def is_admin(user):
    """Check if a user has admin or superuser privileges through role"""
    if hasattr(user, 'profile'):
        return user.profile.role == 'admin' # replace with group check instead
    return user.is_superuser

# Redirect to login if not authenticated
def login_view(request):
    """
    Display and process the login page
    """
    # If user is already authenticated, redirect to home
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'bcit_accreditation/bcit_accred_login.html')

def login_user(request):
    """
    Login form post request using django authentication and hashing
    """
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, "Invalid username or password.")
            return render(request, "bcit_accreditation/bcit_accred_login.html")
    return render(request, "bcit_accreditation/bcit_accred_login.html")

def register_view(request):
    """
    Display and process the register page
    """
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'bcit_accreditation/bcit_accred_register.html')

def register_user(request):
    """
    Register post request, handling logic and adding to database
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        # Handling both entered passwords
        if password1 != password2:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'bcit_accreditation/bcit_accred_register.html')
        
        # Use django password validator
        try:
            # Uses all of django's password validators to check password
            validate_password(password1)
        except ValidationError as e:
            for error in e:
                messages.error(request, error)
            return render(request, "bcit_accreditation/bcit_accred_register.html")

        # Handling username already existing
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already taken.')
            return render(request, 'bcit_accreditation/bcit_accred_register.html')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
            return render(request, 'bcit_accreditation/bcit_accred_register.html')

        # Create user and save it in the database
        user = User.objects.create_user(username=username, email=email, password=password1)
        Faculty.objects.create(user=user, last_uploaded=None)
        user.save()
        login(request, user)
        return redirect('home')

    return render(request, 'bcit_accreditation/bcit_accred_register.html')

def logout_user(request):
    """
    Logout post request
    """
    logout(request)
    return redirect('login')


# ALL FOLLOWING VIEWS MUST HAVE LOGIN_REQUIRED DECORATOR

@login_required
def home_view(request):
    """
    Display the unified home page (requires authentication)
    """
    # Simple context - no additional data needed for the card-based homepage
    context = {}
    
    return render(request, 'bcit_accreditation/bcit_accred_home.html', context)

@login_required
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    """
    Display the admin dashboard (admin only)
    """
    # Get all users except the current admin
    users = User.objects.exclude(id=request.user.id).order_by('username')
    
    # For each user, get their last upload date
    for user in users:
        try:
            # Assuming we have a model that tracks uploads with a user foreign key and a date field
            last_upload = Faculty.objects.filter(user=user).order_by('-last_uploaded').first()
            if last_upload and last_upload.last_uploaded:
                user.last_upload = last_upload.last_uploaded.strftime('%Y-%m-%d')
            else:
                user.last_upload = None
        except Exception as e:
            user.last_upload = None
    
    # Get flattened data from all tables for the database explorer
    # Call the function to get the actual data
    database_entries = get_flattened_data()
    
    # Take only the first 10 entries for initial display
    # The rest will be loaded via API pagination
    display_entries = database_entries[:10] if database_entries else []
    
    context = {
        'users': users,
        'database_entries': display_entries,
        'total_courses': len(set(entry.get('course') for entry in database_entries if entry.get('course'))),
        'total_faculty': len(set(f"{entry.get('instr_first_name')} {entry.get('instr_last_name')}" 
                             for entry in database_entries 
                             if entry.get('instr_first_name') and entry.get('instr_last_name'))),
        'activities': [
            {'date': '2023-07-15', 'user': 'johndoe', 'action': 'Uploaded course data'},
            {'date': '2023-07-14', 'user': 'janedoe', 'action': 'Updated faculty records'},
            {'date': '2023-07-10', 'user': 'admin', 'action': 'System backup'}
        ]
    }
    return render(request, 'bcit_accreditation/bcit_accred_admin.html', context)

@login_required
def csv_upload_view(request):
    """
    Handle CSV or XLSX file upload and extract student data.
    """
    if request.method == 'POST' and request.FILES.get('csv_file'):
        uploaded_file = request.FILES['csv_file']
        file_name = uploaded_file.name.lower()

        try:
            # Convert uploaded file to rows: list of lists
            if file_name.endswith('.csv'):
                file_content = uploaded_file.read().decode('utf-8')
                csvfile = io.StringIO(file_content)
                rows = list(csv.reader(csvfile))

            elif file_name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                sheet = wb.active
                csv_buffer = io.StringIO()
                csv_writer = csv.writer(csv_buffer)
                for row in sheet.iter_rows(values_only=True):
                    csv_writer.writerow(row)
                csv_buffer.seek(0)
                rows = list(csv.reader(csv_buffer))

            else:
                return JsonResponse({'success': False, 'message': 'Unsupported file type. Please upload a .csv or .xlsx file.'})

            # Extract data starting from row 4 (index 3)
            extracted_data = []

            for row in rows[3:]:
                try:
                    student_id = row[1].strip() if row[1] else None
                    gai_score_raw = str(row[2]).strip() if row[2] else None

                    if student_id and len(student_id) == 9:
                        try:
                            gai_score = int(gai_score_raw)
                            extracted_data.append((student_id, gai_score))
                        except ValueError:
                            continue 
                except IndexError:
                    continue  # Skip incomplete rows

            return JsonResponse({
                'success': True,
                'message': 'File uploaded and data extracted.',
                'file_name': uploaded_file.name,
                'file_size': uploaded_file.size,
                'extracted': extracted_data
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})

    return render(request, 'bcit_accreditation/csv_upload.html')

def download_student_assessment_view(request):
    file_path = os.path.join(settings.BASE_DIR, 'references', 'Individual Student Assessment Data.xlsx')

    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='Individual Student Assessment Data.xlsx')

def export_view(request):
    # Data fetching logic goes here: (data = ...)
    data = [
        [f"val{i + 1}_1" for i in range(25)],  # instance 1
        [f"val{i + 1}_2" for i in range(25)],  # instance 2
        [f"val{i + 1}_3" for i in range(25)],  # instance 3
        [f"val{i + 1}_4" for i in range(25)],  # instance 4
        [f"val{i + 1}_5" for i in range(25)],  # instance 5
        [f"val{i + 1}_6" for i in range(25)],  # instance 6
        [f"val{i + 1}_7" for i in range(25)],  # instance 7
        [f"val{i + 1}_8" for i in range(25)],  # instance 8
        [f"val{i + 1}_9" for i in range(25)],  # instance 9
        [f"val{i + 1}_10" for i in range(25)]  # instance 10
    ]

    field_names = [
        "Field 1", "Field 2", "Field 3", "Field 4", "Field 5",
        "Field 6", "Field 7", "Field 8", "Field 9", "Field 10",
        "Field 11", "Field 12", "Field 13", "Field 14", "Field 15",
        "Field 16", "Field 17", "Field 18", "Field 19", "Field 20",
        "Field 21", "Field 22", "Field 23", "Field 24", "Field 25"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"
    ws.append(field_names)

    for row in data:
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=export.xlsx'
    wb.save(response)
    return response

@login_required
def form_step1_view(request):
    """
    Display the first step of the form
    """
    return render(request, 'bcit_accreditation/bcit_accred_f1_course_and_ga.html')

@login_required
def form_step2_view(request):
    """
    Display the second step of the form
    """
    return render(request, 'bcit_accreditation/bcit_accred_f2_info_about_assessment.html')

@login_required
def form_step3_view(request):
    """
    Display the third step of the form (comments & confirmation)
    """
    return render(request, 'bcit_accreditation/bcit_accred_f3_comment_and_confirmation.html')

@login_required
def form_success_view(request):
    """
    Display the success page after form submission
    """
    return render(request, 'bcit_accreditation/bcit_accred_f4_successful_upload.html')

@login_required
def form_submit_view(request):
    """
    Process the form submission
    """
    if request.method == 'POST':
        try:

            # Initial file check
            if 'csv_file' not in request.FILES:
                return JsonResponse({'success': False, 'message': 'No CSV file uploaded'})

            # Get data as list of tuples from function in utils.py
            extracted_data, error = read_csv(request.FILES['csv_file'])

            if error:
                return JsonResponse({'success': False, 'message': error})
            
            if not extracted_data:
                return JsonResponse({'success': False, 'message': 'No valid student data found in uploaded file'})

            program = request.POST.get("program")
            course = request.POST.get("course")
            term = request.POST.get("academicTerm")
            prog_term = int(request.POST.get("programTerm"))
            instr_first_name = request.POST.get("facultyFirstName1")
            instr_last_name = request.POST.get("facultyFirstName2")
            ga = request.POST.get("graduateAttribute")
            gai = request.POST.get("graduateAttributeIndicator")
            instr_level = request.POST.get("instructionalLevel")
            alignment = request.POST.get("assessmentAlignment")
            clos = request.POST.get("courseLearningOutcomes")
            assess_type = request.POST.get("assessmentType")
            assess_weight = float(request.POST.get("courseWeighting"))
            assess_max = request.POST.get("assessmentMaxScore")
            total_score = float(request.POST.get("assessmentTotalScore"))
            question_max = request.POST.get("gaiMaxPoints")
            assess_title = request.POST.get("assessmentTitle")
            assess_descript = request.POST.get("assessmentDescription")
            quest_text = request.POST.get("questionText")
            instr_comments = request.POST.get("assessmentComments")

            for student_id, gai_score in extracted_data:
                res = upload_data(
                    program=program,
                    course=course,
                    term=term,
                    prog_term=prog_term,
                    instr_first_name=instr_first_name,
                    instr_last_name=instr_last_name,
                    ga=ga,
                    gai=gai,
                    instr_level=instr_level,
                    alignment=alignment,
                    clos=clos,
                    assess_type=assess_type,
                    assess_weight=assess_weight,
                    assess_max=int(assess_max),
                    total_score=total_score,
                    question_max=int(question_max),
                    assess_title=assess_title,
                    assess_descript=assess_descript,
                    quest_text=quest_text,
                    instr_comments=instr_comments,
                    student_id=student_id,
                    gai_score=float(gai_score))
                
                # if res.get("success") == False:
                #     print(res.get("message"))
                #     raise Exception
            
            # Change last_updated field
            faculty = Faculty.objects.get(user=request.user)
            faculty.last_uploaded = datetime.now()
            faculty.save()
            
            # Return a success
            print("request: ", request)
            return JsonResponse({
                'success': True,
                'message': 'Data saved successfully!'
            })
        
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({
                'success': False, 
                'message': f'Error saving data: {str(e)}'
            })
    
    # If not POST, redirect to form step 1
    return redirect('form_step1')

# @login_required
# def analytics_view(request):
#     """
#     Display analytics dashboard
#     """
#     return render(request, 'bcit_accreditation/analytics.html')

@login_required
def analysis_view(request):
    """
    Display the new analytics/analysis dashboard with Tableau visualizations
    """
    return render(request, 'bcit_accreditation/bcit_accred_analysis.html')

@login_required
@user_passes_test(is_admin)
def api_data_view(request, table_name):
    """API endpoint for fetching paginated data for the admin dashboard"""
    # Get pagination parameters
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 10))
    sort_by = request.GET.get('sort_by', 'id')
    sort_order = request.GET.get('sort_order', 'asc')
    
    # Calculate offsets for pagination
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    
    try:
        # Get all data using the flattened data function
        all_data = get_flattened_data()
        total_records = len(all_data)
        
        # Custom sorting function to handle different data types
        def get_sort_key(item):
            value = item.get(sort_by, '')
            # Handle None values
            if value is None:
                return '' if sort_order == 'asc' else 'zzzzzzzzzz'  # Empty string for asc, high value for desc
            return value
        
        # Sort the data
        all_data = sorted(all_data, key=get_sort_key, reverse=(sort_order == 'desc'))
        
        # Paginate the data
        paginated_data = all_data[start_idx:end_idx]
        
        # Calculate pagination details
        total_pages = (total_records + page_size - 1) // page_size
        start_record = start_idx + 1 if paginated_data else 0
        end_record = start_idx + len(paginated_data)
        
        # Prepare the response
        response_data = {
            'results': paginated_data,
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'page_size': page_size,
                'total_records': total_records,
                'start_record': start_record,
                'end_record': end_record
            }
        }
        
        return JsonResponse(response_data)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
